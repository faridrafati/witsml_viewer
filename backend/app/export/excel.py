"""XLSX export of curve data via openpyxl.

`build_curves_xlsx` writes a single worksheet: a leading index column (time or
depth) followed by one column per mnemonic. The header row carries the mnemonic
with its unit in parentheses (e.g. ``ROP (m/h)``) so the spreadsheet is
self-describing. Rows are assembled upstream into a flat ``list[dict]`` keyed by
the index label and each mnemonic, so this writer stays format-agnostic.

`assemble_rows` is a convenience helper that pivots the per-mnemonic sample
lists returned by the warm store (or Postgres history) into row dicts aligned on
their shared index. Curves rarely share an exact index value, so alignment is by
the union of all indices seen across the requested mnemonics.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from app.domain.models import CurveSample

# A row index is either a depth (float) or an instant (datetime).
RowIndex = float | datetime


def _index_to_cell(index: RowIndex) -> float | datetime:
    """Normalize an index for an Excel cell (datetimes pass through naive)."""
    if isinstance(index, datetime):
        # openpyxl cannot serialise tz-aware datetimes; drop tzinfo (values are
        # UTC throughout the ingest path) so the cell renders as a date-time.
        return index.replace(tzinfo=None)
    return float(index)


def assemble_rows(
    curves: dict[str, list[CurveSample]],
    mnemonics: list[str],
    index_label: str,
) -> tuple[list[str], list[dict], dict[str, str | None]]:
    """Pivot per-mnemonic sample lists into index-aligned row dicts.

    Returns ``(columns, rows, units)`` where:
      * ``columns`` is ``[index_label, *mnemonics_present]`` preserving the
        requested mnemonic order, restricted to mnemonics that have data;
      * ``rows`` is a list of dicts keyed by ``index_label`` and each present
        mnemonic (missing cells are simply absent / None);
      * ``units`` maps each present mnemonic to its uom (first non-null seen).
    """
    present: list[str] = [m for m in mnemonics if curves.get(m)]
    if not present:
        # Fall back to whatever the store returned (e.g. caller passed []).
        present = [m for m, s in curves.items() if s]

    units: dict[str, str | None] = {}
    # index -> {mnemonic: value}
    by_index: dict[RowIndex, dict[str, float | str | None]] = {}
    order: list[RowIndex] = []

    for mnem in present:
        samples = curves.get(mnem) or []
        if mnem not in units:
            units[mnem] = next((s.uom for s in samples if s.uom), None)
        for s in samples:
            cell = by_index.get(s.index)
            if cell is None:
                cell = {}
                by_index[s.index] = cell
                order.append(s.index)
            cell[mnem] = s.value if s.value is not None else s.text

    order.sort(key=lambda ix: (ix.timestamp() if isinstance(ix, datetime) else ix))

    rows: list[dict] = []
    for ix in order:
        row: dict = {index_label: ix}
        row.update(by_index[ix])
        rows.append(row)

    return [index_label, *present], rows, units


def build_curves_xlsx(
    well_uid: str,
    columns: list[str],
    rows: list[dict],
    index_label: str,
    *,
    units: dict[str, str | None] | None = None,
) -> bytes:
    """Render curve rows to .xlsx bytes.

    ``columns`` is the ordered list of keys (first must be ``index_label``).
    ``rows`` are dicts keyed by those columns. ``units`` (optional) maps a
    column name to its uom; when given the header cell becomes ``name (uom)``.
    """
    units = units or {}
    wb = Workbook()
    ws = wb.active
    ws.title = (well_uid or "curves")[:31] or "curves"

    # Header row: index column first, then one column per mnemonic with unit.
    header: list[str] = []
    for col in columns:
        if col == index_label:
            header.append(index_label)
            continue
        uom = units.get(col)
        header.append(f"{col} ({uom})" if uom else col)
    ws.append(header)

    is_time_index = any(isinstance(r.get(index_label), datetime) for r in rows)

    for r in rows:
        out_row: list[float | str | datetime | None] = []
        for col in columns:
            val = r.get(col)
            if col == index_label and val is not None:
                out_row.append(_index_to_cell(val))
            else:
                out_row.append(val)
        ws.append(out_row)

    # Format the index column for readability.
    if is_time_index:
        for cell in ws[get_column_letter(1)][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # Modest fixed widths so the export opens legibly.
    for idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 22 if idx == 1 else 16

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
